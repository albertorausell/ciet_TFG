from django.shortcuts import render, redirect
from common_app.models import activity, capability_learner, evaluation, exercise, learner_profile, organization, question, answer, trainer_profile, training_technique, objective, capability, capability_objective, content
from .forms import Excel_form, Text, Image, Video, Document, Link, Game, Contents, ObjectiveForm, CapabilityName, CapabilityDesc, CapabilityLearners, CapabilityObjectives
import json
import math
from django.http import HttpResponse
from django.utils.safestring import SafeString
from django.contrib.auth.decorators import login_required, user_passes_test
from openpyxl import load_workbook
# Create your views here.

# ctrl+shift+p > Preferences: Configure Language Specific Settings > Python
# {
#    "python.linting.pylintArgs": [
#        "--load-plugins=pylint_django"
#    ],
#
#    "[python]": {
#
#    }
# }
# https://stackoverflow.com/questions/45135263/class-has-no-objects-member


def is_trainer(user):
    try:
        return trainer_profile.objects.filter(user=user.pk)[0].rol.isTrainer
    except:
        return False


@login_required(redirect_field_name=None)
@user_passes_test(is_trainer, login_url='/learner/capabilities', redirect_field_name=None)
def capabilities(request):
    dictionary = create_base_dictionary(request)

    cap = capability.objects.filter(
        trainer=dictionary['trainer'].pk, organization=dictionary['org'].pk).order_by('name')

    cap_name = []
    cap_obj = []
    cap_stk = []
    cap_act = []
    cap_id = []

    for i in cap:
        #
        cap_name.append(i.name)
        #
        objectives_chain = ''
        objectives = capability_objective.objects.filter(capability=i.pk)
        for j in objectives:
            the_objective = objective.objects.filter(
                pk=j.objective.pk)[0]
            objectives_chain += the_objective.name + ', '
        cap_obj.append(objectives_chain[:-2])
        #
        stakeholders = i.stakeholders.all()
        stakeholders_chain = ''
        for j in stakeholders:
            stakeholders_chain += j.name + ', '
        cap_stk.append(stakeholders_chain[:-2])
        #
        cap_act.append(i.active)
        #
        cap_id.append(i.pk)

    rows = zip(cap_name, cap_obj, cap_stk, cap_act, cap_id)

    dictionary.update({
        'table': {
            'columns': ['Name', 'Objectives', 'Learners', 'Active', 'Actions'],
            'rows': rows
        }
    })

    return render(request, 'trainer_app/templates/capabilities.html', dictionary)


@login_required(redirect_field_name=None)
@user_passes_test(is_trainer, login_url='/learner/capabilities', redirect_field_name=None)
def changeOrganization(request, pos):
    trainer = trainer_profile.objects.get(user=request.user.pk)
    trainer.actual_organization_pos = pos
    trainer.save()
    return redirect('capabilities_trainer')


@login_required(redirect_field_name=None)
@user_passes_test(is_trainer, login_url='/learner/capabilities', redirect_field_name=None)
def setActive(request, id):
    cap = capability.objects.get(pk=id)
    if cap.active == True:
        cap.active = False
        cap.save()
    else:
        cap.active = True
        cap.save()
    return HttpResponse(status=200)


@login_required(redirect_field_name=None)
@user_passes_test(is_trainer, login_url='/learner/capabilities', redirect_field_name=None)
def setPending(request, id):
    ex = exercise.objects.get(pk=id)
    if ex.leftPending == True:
        ex.leftPending = False
        ex.save()
    else:
        ex.leftPending = True
        ex.save()
    return HttpResponse(status=200)


@login_required(redirect_field_name=None)
@user_passes_test(is_trainer, login_url='/learner/capabilities', redirect_field_name=None)
def cap_name(request, id_str):
    id = int(id_str)
    dictionary = create_base_dictionary(request)

    if request.method == 'POST':
        form = CapabilityName(request.POST)
        if form.is_valid():
            cap = capability.objects.get(pk=id)
            cap.name = form.cleaned_data.get('cap_name')
            cap.save()
            return redirect('description_and_image', id=id)
    else:
        if id == -1:
            cap = capability()
            cap.trainer = dictionary['trainer']
            cap.organization = dictionary['org']
            cap.save()
            form = CapabilityName()
            dictionary.update({
                'cap_id': cap.pk,
                'form': form,
            })
        else:
            cap = capability.objects.get(pk=id)
            form = CapabilityName(initial={'cap_name': cap.name})
            dictionary.update({
                'cap_id': cap.pk,
                'form': form,
            })

        return render(request, 'capabilities/create_name.html', dictionary)

    return HttpResponse(status=204)


@login_required(redirect_field_name=None)
@user_passes_test(is_trainer, login_url='/learner/capabilities', redirect_field_name=None)
def cap_desc(request, id):
    dictionary = create_base_dictionary(request)
    cap = capability.objects.get(pk=id)

    if request.method == 'POST':
        form = CapabilityDesc(request.POST, request.FILES)
        if form.is_valid():
            cap.description = form.cleaned_data.get('cap_desc')
            cap.image = form.cleaned_data.get('cap_image')
            cap.save()
            return redirect('learners', id=id)
        else:
            if (cap.image):
                cap.description = form.cleaned_data.get('cap_desc')
                cap.save()
                return redirect('learners', id=id)
    else:
        form = CapabilityDesc(
            initial={'cap_desc': cap.description, 'cap_image': cap.image})

        dictionary.update({
            'cap_id': cap.pk, 'form': form})

        return render(request, 'capabilities/create_description.html', dictionary)

    return HttpResponse(status=204)


@login_required(redirect_field_name=None)
@user_passes_test(is_trainer, login_url='/learner/capabilities', redirect_field_name=None)
def cap_learners(request, id):
    dictionary = create_base_dictionary(request)
    cap = capability.objects.get(pk=id)

    if request.method == 'POST':
        form = CapabilityLearners(request.POST)
        if form.is_valid():
            cap.stakeholders.set(form.cleaned_data.get("stakeholders"))
            cap.save()
            return redirect('objectives', id=id)
    else:
        form = CapabilityLearners(instance=cap)
        dictionary.update({
            'cap_id': cap.pk, 'form': form
        })
        return render(request, 'capabilities/create_learners.html', dictionary)

    return HttpResponse(status=204)


@login_required(redirect_field_name=None)
@user_passes_test(is_trainer, login_url='/learner/capabilities', redirect_field_name=None)
def cap_objectives(request, id):
    dictionary = create_base_dictionary(request)
    cap = capability.objects.get(pk=id)

    if request.method == 'POST':
        form = CapabilityObjectives(request.POST)
        if form.is_valid():
            cap.objectives.set(form.cleaned_data.get("objectives"))
            cap.save()
            return redirect('contents', id=id, obj_pos=0, cont_pos=0, view=0, ev_pos=0)
    else:
        form = CapabilityObjectives(instance=cap)
        dictionary.update({
            'cap_id': cap.pk
        })

        obj_act = []
        obj_ph = []
        objectives = objective.objects.filter(
            organization=dictionary['org'].pk)
        for ob in objectives:
            act_chain = ''
            ph_chain = ''
            ob_activities = ob.activities.all()
            for act in ob_activities:
                act_chain += act.name + ', '
                try:
                    ph_chain.index(act.phase.name)
                except:
                    ph_chain += act.phase.name + ', '
            obj_act.append(act_chain[:-2])
            obj_ph.append(ph_chain[:-2])

        dictionary.update({
            'obj_act': obj_act,
            'obj_ph': obj_ph,
            'form': form
        })
        return render(request, 'capabilities/create_objectives.html', dictionary)

    return HttpResponse(status=204)


@login_required(redirect_field_name=None)
@user_passes_test(is_trainer, login_url='/learner/capabilities', redirect_field_name=None)
def cap_contents(request, id, obj_pos, cont_pos, view, ev_pos):
    dictionary = create_base_dictionary(request)
    cap = capability.objects.get(pk=id)

    if request.method == 'POST':
        post = request.POST
        value = post['action']
        if '_c' in value:
            idObj = value[:-2]
            form = Contents(request.POST)
            if form.is_valid():
                objective = capability_objective.objects.filter(
                    capability=cap.pk, objective=idObj)
                cont = content()
                cont.name = form.cleaned_data.get("name")
                cont.dimension = form.cleaned_data.get("dimension")
                cont.capability_objective = objective[0]
                cont.save()
                return redirect('contents', id=id, obj_pos=int(request.POST['objMant']), cont_pos=int(request.POST['contMant']), view=0, ev_pos=0)
        if '&txt' in value:
            idCon = value[:-4]
            form = Text(request.POST)
            if form.is_valid():
                txt = form.save(commit=False)
                txt.order = getOrder(idCon)
                txt.type = 'txt'
                txt.content = content.objects.get(pk=idCon)
                txt.save()
                return redirect('contents', id=id, obj_pos=int(request.POST['objMant']), cont_pos=int(request.POST['contMant']), view=0, ev_pos=0)

        if '&img' in value:
            idCon = value[:-4]
            form = Image(request.POST, request.FILES)
            if form.is_valid():
                img = form.save(commit=False)
                img.order = getOrder(idCon)
                img.type = 'img'
                img.content = content.objects.get(pk=idCon)
                img.save()
                return redirect('contents', id=id, obj_pos=int(request.POST['objMant']), cont_pos=int(request.POST['contMant']), view=0, ev_pos=0)

        if '&vid' in value:
            idCon = value[:-4]
            form = Video(request.POST, request.FILES)
            if form.is_valid():
                vid = form.save(commit=False)
                vid.order = getOrder(idCon)
                vid.type = 'vid'
                vid.content = content.objects.get(pk=idCon)
                vid.save()
                return redirect('contents', id=id, obj_pos=int(request.POST['objMant']), cont_pos=int(request.POST['contMant']), view=0, ev_pos=0)

        if '&doc' in value:
            idCon = value[:-4]
            form = Document(request.POST, request.FILES)
            if form.is_valid():
                doc = form.save(commit=False)
                doc.order = getOrder(idCon)
                doc.type = 'doc'
                doc.content = content.objects.get(pk=idCon)
                doc.save()
                return redirect('contents', id=id, obj_pos=int(request.POST['objMant']), cont_pos=int(request.POST['contMant']), view=0, ev_pos=0)

        if '&lnk' in value:
            idCon = value[:-4]
            form = Link(request.POST)
            if form.is_valid():
                lnk = form.save(commit=False)
                lnk.order = getOrder(idCon)
                lnk.type = 'lnk'
                lnk.content = content.objects.get(pk=idCon)
                lnk.save()
                return redirect('contents', id=id, obj_pos=int(request.POST['objMant']), cont_pos=int(request.POST['contMant']), view=0, ev_pos=0)

        if '&gme' in value:
            idCon = value[:-4]
            form = Game(request.POST)
            if form.is_valid():
                gme = form.save(commit=False)
                gme.order = getOrder(idCon)
                gme.type = 'gme'
                gme.content = content.objects.get(pk=idCon)
                gme.save()
                return redirect('contents', id=id, obj_pos=int(request.POST['objMant']), cont_pos=int(request.POST['contMant']), view=0, ev_pos=0)

        if 'addQuestion' in value:
            addQuestionFromPost(post, cap)
            return redirect('contents', id=id, obj_pos=int(request.POST['objMant']), cont_pos=int(request.POST['contMant']), view=1, ev_pos=int(request.POST['evMant']))

    else:
        # CONTENTS
        objs = cap.objectives.all()
        dictionary.update({
            'obj_pos': obj_pos,
            'cont_pos': cont_pos,
            'view': view,
            'ev_pos': ev_pos,
            'cap_name': cap.name,
            'cap_id': cap.pk,
            'cap_obj': objs,
            'contentsForm': Contents(),
            'componentsForms': {
                'text': Text(),
                'image': Image(),
                'video': Video(),
                'document': Document(),
                'link': Link(),
                'game': Game(),
            },
        })
        obj_cont = []
        cont_tech = []
        for obj in objs:
            cap_obj = capability_objective.objects.filter(
                capability=cap.pk, objective=obj.pk)
            contents = content.objects.filter(
                capability_objective=cap_obj[0].pk)
            obj_cont.append({
                'obj_pk': obj.pk,
                'cap_obj_pk': cap_obj[0].pk,
                'name': obj.name,
                'contents': contents
            })
            for cont in contents:
                media_comps = training_technique.objects.filter(
                    content=cont.pk
                ).select_subclasses()
                cont_tech.append({
                    'cont': cont,
                    'media_comps': orderComps(media_comps),
                })
        dictionary.update({
            'obj_cont': obj_cont,
            'cont_tech': cont_tech,
        })

        # EVALUATIONS
        exs = []
        exercises = exercise.objects.filter(capability=cap)
        for ex in exercises:
            qss = []
            questions = question.objects.filter(exercise=ex)
            for ques in questions:
                answers = answer.objects.filter(question=ques)
                qss.append({
                    'question': ques,
                    'answers': answers
                })
            exs.append({
                'exercise': ex,
                'questions': qss
            })
        dictionary.update({
            'exercises': exs
        })

        return render(request, 'capabilities/create_contents.html', dictionary)

    return HttpResponse(status=204)


def getOrder(idCon):
    components = training_technique.objects.filter(
        content=idCon
    ).select_subclasses()
    max = 0
    for comp in components:
        if comp.order > max:
            max = comp.order

    return max + 1


@login_required(redirect_field_name=None)
@user_passes_test(is_trainer, login_url='/learner/capabilities', redirect_field_name=None)
def editQuestionReq(request):
    deleteQuestion(request.POST['idQuestion'])
    cap = capability.objects.get(pk=request.POST['idCap'])
    addQuestionFromPost(request.POST, cap)
    return redirect('contents', id=request.POST['idCap'], obj_pos=int(request.POST['objMant']), cont_pos=int(request.POST['contMant']), view=1, ev_pos=int(request.POST['evMant']))


def addQuestionFromPost(post, cap):
    ex_list = exercise.objects.filter(
        capability=cap.pk, referenceId=post['idRefQues'], scope=post['typeRefQues'])
    if len(ex_list) > 0:
        quest = question()
        quest.question = post['question']
        quest.exercise = ex_list[0]
        quest.save()
        i = 0
        length = int(post['length'])
        while i < length:
            name = 'answer' + str(i)
            correct = 'correct' + str(i)
            if len(post[name]) > 0:
                ans = answer()
                ans.answer = post[name]
                if correct in post:
                    if post[correct] == 'on':
                        ans.isCorrect = True
                    else:
                        ans.isCorrect = False
                else:
                    ans.isCorrect = False
                ans.question = quest
                ans.save()
            i += 1
    else:
        ex = exercise()
        ex.scope = post['typeRefQues']
        ex.referenceId = post['idRefQues']
        ex.capability = cap
        ex.save()
        quest = question()
        quest.question = post['question']
        quest.exercise = ex
        quest.save()
        i = 0
        length = int(post['length'])
        while i < length:
            name = 'answer' + str(i)
            correct = 'correct' + str(i)
            if len(post[name]) > 0:
                ans = answer()
                ans.answer = post[name]
                if correct in post:
                    if post[correct] == 'on':
                        ans.isCorrect = True
                    else:
                        ans.isCorrect = False
                else:
                    ans.isCorrect = False
                ans.question = quest
                ans.save()
            i += 1


@login_required(redirect_field_name=None)
@user_passes_test(is_trainer, login_url='/learner/capabilities', redirect_field_name=None)
def deleteQuestionReq(request):
    deleteQuestion(request.POST['idQuestion'])
    return redirect('contents', id=request.POST['idCap'], obj_pos=0, cont_pos=0, view=1, ev_pos=int(request.POST['evMant']))


def deleteQuestion(id):
    question_to_delete = question.objects.get(pk=id)
    answers_to_delete = answer.objects.filter(question=id)
    for ans in answers_to_delete:
        ans.delete()
    question_to_delete.delete()


@login_required(redirect_field_name=None)
@user_passes_test(is_trainer, login_url='/learner/capabilities', redirect_field_name=None)
def deleteContentReq(request):
    deleteContent(request.POST['idCont'])
    return redirect('contents', id=request.POST['idCap'], obj_pos=request.POST['objMant'], cont_pos=request.POST['contMant'], view=0, ev_pos=0)


def deleteContent(id):
    cont_to_delete = content.objects.get(pk=id)
    comps_to_delete = training_technique.objects.filter(content=id)
    for comp in comps_to_delete:
        deleteComponent(comp.pk)
    cont_to_delete.delete()


@login_required(redirect_field_name=None)
@user_passes_test(is_trainer, login_url='/learner/capabilities', redirect_field_name=None)
def deleteComponentReq(request):
    deleteComponent(request.POST['compId'])
    return redirect('contents', id=request.POST['idCap'], obj_pos=request.POST['objPos'], cont_pos=request.POST['contPos'], view=0, ev_pos=0)


def deleteComponent(id):
    tt_to_delete = training_technique.objects.filter(pk=id)
    comp_to_delete = tt_to_delete.select_subclasses()[0]
    tt_to_delete = tt_to_delete[0]
    tt_to_delete.delete()
    comp_to_delete.delete()


@login_required(redirect_field_name=None)
@user_passes_test(is_trainer, login_url='/learner/capabilities', redirect_field_name=None)
def objectives(request):
    dictionary = create_base_dictionary(request)

    objectivesRows = objective.objects.filter(
        organization=dictionary['org'].pk).order_by('name')
    dictionary.update({
        'table': {
            'columns': ['Objective', 'Actions'],
            'rows': objectivesRows
        }
    })

    objective_form = ObjectiveForm()
    excel_form = Excel_form()
    dictionary.update({
        'objectivesForm': objective_form,
        'excel_form': excel_form
    })

    if request.method == 'POST':
        form = ObjectiveForm(request.POST)
        if form.is_valid():
            form.save()

    return render(request, 'objectives.html', dictionary)


@login_required(redirect_field_name=None)
@user_passes_test(is_trainer, login_url='/learner/capabilities', redirect_field_name=None)
def import_objectives(request):
    dictionary = create_base_dictionary(request)
    if request.method == 'POST':
        form = Excel_form(request.POST, request.FILES)
        if form.is_valid():
            doc = form.save(commit=True)
            path = doc.document.path
            wb = load_workbook(filename=path)
            worksheet = wb.worksheets[1]
            objs = objective.objects.filter(organization=dictionary['org'])
            objs = list(map(lambda x: x.name, objs))
            acts_list = list(activity.objects.all())
            acts_name_list = list(
                map(lambda x: x.name.lower().strip(), acts_list))
            i = 4  # when title "TRAINING CONTENT" is reached
            while i <= worksheet.max_row:
                cell = worksheet['D' + str(i)].value
                if cell not in objs and cell != None:
                    acts = worksheet['F' + str(i)].value.lower().strip()
                    acts = acts.split('/')
                    acts_to_add = []
                    for act in acts:
                        if act not in acts_name_list:
                            doc.delete()
                            return HttpResponse('Activity "' + worksheet['F' + str(i)].value + '" in cell "F' + str(i) + '" not found. Go back and try again.')
                        acts_to_add.append(
                            acts_list[acts_name_list.index(act)])
                    new_objective = objective(
                        name=cell, organization=dictionary['org'])
                    new_objective.save()
                    new_objective.activities.set(acts_to_add)
                    new_objective.save()
                    objs.append(cell)
                i += 1

    return redirect('objectives_trainer')


@login_required(redirect_field_name=None)
@user_passes_test(is_trainer, login_url='/learner/capabilities', redirect_field_name=None)
def statistics(request):
    dictionary = create_base_dictionary(request)
    # first chart
    caps = capability.objects.filter(organization=dictionary['org'])
    objectives = []
    objs_names = []
    for cap in caps:
        cap_learns_reached = capability_learner.objects.filter(
            capability=cap.pk, pending=False)
        cap_learns_reached = list(
            map(lambda x: x.learner_profile, cap_learns_reached))
        stks = cap.stakeholders.all()
        cap_learns_total = []
        for stk in stks:
            cap_learns = learner_profile.objects.filter(
                organization=dictionary['org'], rol=stk.pk)
            cap_learns_total = list(set(cap_learns_total) | set(cap_learns))

        rols = {}
        for cap_learn in cap_learns_total:
            if cap_learn in cap_learns_reached:
                if cap_learn.rol.name in rols:
                    rols[cap_learn.rol.name] = {
                        'reached': rols[cap_learn.rol.name]['reached'] + 1,
                        'total': rols[cap_learn.rol.name]['total'] + 1
                    }
                else:
                    rols[cap_learn.rol.name] = {
                        'reached': 1,
                        'total': 1
                    }
            else:
                if cap_learn.rol.name in rols:
                    rols[cap_learn.rol.name] = {
                        'reached': rols[cap_learn.rol.name]['reached'],
                        'total': rols[cap_learn.rol.name]['total'] + 1
                    }
                else:
                    rols[cap_learn.rol.name] = {
                        'reached': 0,
                        'total': 1
                    }
        objs = cap.objectives.all()
        for obj in objs:
            if obj.name in objs_names:
                item = objectives[objs_names.index(obj.name)]
                keys = rols.keys()
                for key in keys:
                    if key in item['rols']:
                        item['rols'][key]['reached'] += rols[key]['reached']
                        item['rols'][key]['total'] += rols[key]['total']
                    else:
                        item['rols'][key] = rols[key]
            else:
                objectives.append({
                    'objective': obj.name,  # safestring¿?
                    'rols': rols
                })
                objs_names.append(obj.name)

    for obj in objectives:
        obj['rols_keys'] = list(obj['rols'].keys())
        obj['rols_keys_length'] = len(obj['rols'].keys())
    set_labels = []
    for obj in objectives:
        for key in obj['rols_keys']:
            if key not in set_labels:
                set_labels.append(key)
    dictionary.update({
        'chart1_sets': set_labels,
        'first_chart': objectives
    })
    # ---------------------------------------------
    # second chart
    data_exercises = {}
    org_exercises = []
    for cap in caps:
        exs = list(exercise.objects.filter(capability=cap.pk))
        org_exercises = org_exercises + exs

    caps_with_exercises = []
    org_learners = learner_profile.objects.filter(
        organization=dictionary['org'])
    for learner in org_learners:
        for ex in org_exercises:
            first_ev = evaluation.objects.filter(
                exercise=ex.pk, learner_profile=learner.pk).order_by('created_at')
            last_ev = evaluation.objects.filter(
                exercise=ex.pk, learner_profile=learner.pk).order_by('-created_at')
            if len(first_ev) > 0:
                if ex.pk in data_exercises:
                    data_exercises[ex.pk]['first_sum'] += first_ev[0].mark
                    data_exercises[ex.pk]['last_sum'] += last_ev[0].mark
                    data_exercises[ex.pk]['num'] += 1
                else:
                    data_exercises[ex.pk] = {
                        'first_sum': first_ev[0].mark,
                        'last_sum': last_ev[0].mark,
                        'num': 1,
                        'cap': ex.capability
                    }
                    if ex.capability not in caps_with_exercises:
                        caps_with_exercises.append(ex.capability)
    res = []
    for cap in caps_with_exercises:
        first_average = 0
        last_average = 0
        total_num = 0
        for data in data_exercises:
            if data_exercises[data]['cap'] == cap:
                first_average += data_exercises[data]['first_sum']
                last_average += data_exercises[data]['last_sum']
                total_num += data_exercises[data]['num']
        res.append({
            'capability_name': cap.name,
            'first_average': truncate(first_average / total_num, 2),
            'last_average': (truncate(last_average / total_num, 2)) - (truncate(first_average / total_num, 2)),
        })
    dictionary.update({
        'second_chart': res
    })

    # classification
    organization_learners = learner_profile.objects.filter(
        organization=dictionary['org']).order_by('-points')
    dictionary.update({
        'classification': organization_learners
    })
    return render(request, 'statistics.html', dictionary)


def truncate(number, digits) -> float:
    stepper = 10.0 ** digits
    return math.trunc(stepper * number) / stepper


@login_required(redirect_field_name=None)
@user_passes_test(is_trainer, login_url='/learner/capabilities', redirect_field_name=None)
def objective_edit(request, id):
    dictionary = create_base_dictionary(request)

    objective_to_edit = objective.objects.get(pk=id)
    objective_form = ObjectiveForm(instance=objective_to_edit)
    dictionary.update({'objetive_to_edit': objective_form})

    if request.method == 'POST':
        form = ObjectiveForm(request.POST, instance=objective_to_edit)
        if form.is_valid():
            acts = form.cleaned_data.get("activities")
            obj = form.save(commit=False)
            obj.activities.set(acts)
            obj.save()
        return redirect('objectives_trainer')

    return render(request, 'objectives/edit_objective.html', dictionary)


def orderComps(comps):
    res = []
    for comp in comps:
        if len(res) == 0:
            res.append(comp)
        else:
            i = 0
            inserted = False
            for element in res:
                if (not inserted and comp.order < element.order):
                    res.insert(i, comp)
                    inserted = True
                i = i + 1
            if (not inserted):
                res.append(comp)

    return res


def create_base_dictionary(request):
    user = request.user
    name_to_display = user.first_name + ' ' + user.last_name
    if name_to_display == ' ':
        name_to_display = user.username

    trainer = trainer_profile.objects.get(user=user.pk)
    organizations = trainer.organizations.all().order_by('pk')

    dictionary = {
        'name': name_to_display,
        'trainer': trainer,
        'org': organizations[trainer.actual_organization_pos],
        'orgs': organizations,
        'host': 'http://127.0.0.1:8000/'
    }
    return dictionary
